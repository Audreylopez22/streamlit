from tools import log_message
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

def create_simple_sheet(workbook):

   log_message("creating simple sheet")
   
   if "Simple Sheet" not in workbook.sheetnames:
        simple_sheet = workbook.create_sheet("Simple Sheet")

        new_columns = ["Name", "Base Salary", "Ongoing Reimbursements",
                       "Total", "Comments", "Crypto Currency"]

        for i, column_name in enumerate(new_columns, start=1):
            new_column_letter = get_column_letter(i)
            simple_sheet[f"{new_column_letter}1"] = column_name

def main(workbook, progress_bar):
    for sheet in workbook:
        create_simple_sheet(workbook)
    
    if progress_bar is not None:
        progress_bar.progress(1.0 / len(sheet.parent.sheetnames))   

    return workbook